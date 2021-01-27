# General
import time
import datetime
from Scraper import Scraper
from random import randrange

# Excel
from openpyxl import load_workbook

# Reading and Writing
workbookToReadWriteName = 'TripAdvisorScrapeURLs.xlsx'
wbReadWrite = load_workbook(workbookToReadWriteName)
readingWorkingFile = wbReadWrite["Sheet1"]

# Writing via Appending
workbookToWriteName = 'TripAdvisorRawData.xlsx'
wbWrite = load_workbook(workbookToWriteName)
wbWritePage = wbWrite.active

# Column numbers
shopNameColumn              = 2
noOfReviewsColumn           = 4
noOfNewReviewsColumn        = 5
dateUpdatedColumn           = 6 
urlColumn                   = 7

# Debug mode
debugMode = False

dateUpdated = datetime.datetime.now().strftime("%d/%m/%Y")

# Step2: Traverse each row in the excel sheet starting from the second row 
for rowno, rowval in enumerate(readingWorkingFile.iter_rows(min_row=2, max_row=readingWorkingFile.max_row), start=2):    
   
    urlOfShop   = readingWorkingFile.cell(row=rowno,column=urlColumn).value
    noOfReviews = 0 if (readingWorkingFile.cell(row=rowno,column=noOfReviewsColumn).value == None) else readingWorkingFile.cell(row=rowno,column=noOfReviewsColumn).value
    nameOfShop  = readingWorkingFile.cell(row=rowno,column=shopNameColumn).value
    
    # Step3.1: If the URL column does not contain https://www.tripadvisor.com.sg/Restaurant_Review- it just means its not a valid url
    if(urlOfShop == None or "https://www.tripadvisor.com.sg/Restaurant_Review-" not in urlOfShop):
        readingWorkingFile.cell(row=rowno, column=noOfReviewsColumn).value    = "-"
        readingWorkingFile.cell(row=rowno, column=noOfNewReviewsColumn).value = "-"
        readingWorkingFile.cell(row=rowno, column=dateUpdatedColumn).value    = "-"
        
    # Step3.2: Only if it contains the starting valid URL 
    else:   
        with Scraper(debug=debugMode) as scraper:
            scraper.setURL(name=nameOfShop, url=urlOfShop)
            totalNewNoOfReviews = scraper.getNewNumberOfReviews()
            newReviewsCount = totalNewNoOfReviews - noOfReviews
   
            if (newReviewsCount != 0):
                listOfNewReviews=[]
                listOfNewReviews = scraper.getReviews(newReviewsCount,dateUpdated,nameOfShop)
                
                for review in listOfNewReviews:
                    wbWritePage.append(review)
                    
                print("Updated %d new review(s) for %s" % (newReviewsCount, nameOfShop))
            
            else:
                print("No new reviews for %s" % (nameOfShop))
                
            readingWorkingFile.cell(row=rowno, column=noOfReviewsColumn).value    = totalNewNoOfReviews
            readingWorkingFile.cell(row=rowno, column=noOfNewReviewsColumn).value = newReviewsCount
            readingWorkingFile.cell(row=rowno, column=dateUpdatedColumn).value    = dateUpdated

            sleepTime = randrange(5,15)
            print("Sleeping for", sleepTime,"seconds before next shop\n")
            time.sleep(sleepTime)
                
        
# Step 8: Save the file or nothing will be saved
wbReadWrite.save(workbookToReadWriteName)
wbWrite.save(workbookToWriteName)