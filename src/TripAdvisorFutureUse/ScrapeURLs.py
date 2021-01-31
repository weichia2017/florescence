# Download latest chromedriver from: https://sites.google.com/a/chromium.org/chromedriver/downloads

# Only for MAC When u run and get this: “chromedriver” cannot be opened because the developer cannot be verified. 
#       Type this in the terminal: xattr -d com.apple.quarantine ./chromedriver

# Selenium Documentation: https://selenium-python.readthedocs.io/
import time
from random import randrange

from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from openpyxl import load_workbook


workbookToReadName = 'TripAdvisorScrapeURLs.xlsx'

# Step1: Load the xlsx file to be read and written
wbRead = load_workbook(workbookToReadName)

# Step2: Open the working file using workbook and sheet names
readingWorkingFile = wbRead["Sheet1"]

# Step3: Load up chromedriver and visit TripAdvisor Main Page
driver = webdriver.Chrome("./chromedriver")
driver.get("https://www.tripadvisor.com.sg")

# Step4: Find for the search bar in TripAdvisor's Main Page
searchBar = driver.find_element_by_xpath("//input[@placeholder='Where to?']")

rowToWriteOn = 0
urlColumn    = 6
shopColumn   = 1

# Checker to update the command prompt at the end
noOfNewShops              = 0
noOfNewShopsInTripAdvisor = 0

# Step5: Read the shop name row by row and enter it into the search bar and collect the URL's
#        of the shops hosted on TripAdvisor
for row in readingWorkingFile.iter_rows():
    rowToWriteOn += 1
    
    shopName     = row[shopColumn].value 
    shopURL      = row[urlColumn].value 
    
    if (shopName != "Name" and shopURL == None):
        noOfNewShops += 1
             
        # Clear the search bar value
        searchBar.clear()
        
        # Send the shop name to the search bar 
        searchBar.send_keys(shopName)

        sleepTime = randrange(1,4)
        print("Sleeping for", sleepTime,"seconds before next shop")
        time.sleep(sleepTime)

        # Get a list of search results
        searchResultsList = driver.find_elements_by_class_name("_1dvyiAq4")

        isShopListedOnTripAdvisor = False

        for searchResult in searchResultsList:
            # Get the URL to be used if any
            url = searchResult.get_attribute('href')
            name = searchResult.text.split("\n")[0]
            
            # Filter 1
            # Since trip advisor does not only have restaurants, filter by restaurant
            if ("Restaurant_Review" in url):
                country = ((url.split("Reviews-")[1]).split("-")[1]).split(".html")[0]
           
                # Filter 2
                # If country not Singapore ignore and continue loop
                if (country != "Singapore"):
                    continue
                
                # So far .replace("-","") only used for Bar-Celona
                if ((name.lower()).replace("-","") == shopName.lower()):
                    isShopListedOnTripAdvisor = True
                    noOfNewShopsInTripAdvisor += 1
                    readingWorkingFile.cell(row=rowToWriteOn, column=urlColumn+1).value = url
                    break
                
        if(isShopListedOnTripAdvisor == False):
            readingWorkingFile.cell(row=rowToWriteOn, column=urlColumn+1).value = "Does Not Exist"    

            
print("\n===============================================")
print("There were %d shop(s) with no URLs." % (noOfNewShops))
print("All shops have been checked for URLs")
if(noOfNewShops != 0 ):
    print("%d/%d shop(s) are available in TripAdvisor" % (noOfNewShopsInTripAdvisor,noOfNewShops))
print("===============================================\n")
        
# Step 6: Save the file or nothing will be saved
wbRead.save(workbookToReadName)

# Step 7: Close the chromedriver
driver.quit()  